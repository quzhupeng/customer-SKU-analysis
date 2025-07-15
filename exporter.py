#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel报告导出模块
Excel Report Export Module

Author: Augment Agent
Date: 2025-07-15
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import logging
from typing import Dict, Any

logger = logging.getLogger(__name__)

class ReportExporter:
    """Excel报告导出器"""
    
    def __init__(self, analysis_data: Dict[str, Any]):
        """
        初始化导出器
        
        Args:
            analysis_data: 分析结果数据
        """
        self.analysis_data = analysis_data
        self.workbook = Workbook()
        
        # 样式定义
        self.header_font = Font(bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_to_excel(self, filepath: str) -> None:
        """
        导出完整的Excel报告
        
        Args:
            filepath: 导出文件路径
        """
        try:
            # 删除默认的工作表
            if 'Sheet' in self.workbook.sheetnames:
                self.workbook.remove(self.workbook['Sheet'])
            
            # 创建四个工作表
            self._create_summary_sheet()
            self._create_charts_sheet()
            self._create_analyzed_data_sheet()
            self._create_raw_data_sheet()
            
            # 保存文件
            self.workbook.save(filepath)
            logger.info(f"Excel报告已导出到: {filepath}")
            
        except Exception as e:
            logger.error(f"导出Excel报告失败: {str(e)}")
            raise
    
    def _create_summary_sheet(self) -> None:
        """创建分析概要工作表"""
        ws = self.workbook.create_sheet("分析概要 (Report)")
        
        # 报告标题
        ws['A1'] = "产品客户价值分析报告"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = self.center_alignment
        ws.merge_cells('A1:E1')
        
        # 分析信息
        analysis_result = self.analysis_data.get('analysis_result', {})
        analysis_type = analysis_result.get('analysis_type', '')
        
        analysis_type_names = {
            'product': '分单品分析',
            'customer': '分客户分析',
            'region': '分地区分析'
        }
        
        ws['A3'] = "分析类型:"
        ws['B3'] = analysis_type_names.get(analysis_type, analysis_type)
        ws['A4'] = "分析时间:"
        ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A5'] = "数据来源:"
        ws['B5'] = self.analysis_data.get('file_info', {}).get('original_filename', '')
        
        # 设置标题样式
        for cell in ['A3', 'A4', 'A5']:
            ws[cell].font = self.header_font
        
        # 四象限分析概要
        quadrant_analysis = analysis_result.get('quadrant_analysis', {})
        quadrant_stats = quadrant_analysis.get('quadrant_stats', {})
        
        ws['A7'] = "四象限分析概要"
        ws['A7'].font = self.title_font
        ws.merge_cells('A7:E7')
        
        # 象限统计表头
        headers = ['象限', '名称', '特征', '数量', '策略建议']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # 象限数据
        for quadrant in [1, 2, 3, 4]:
            if quadrant in quadrant_stats:
                stats = quadrant_stats[quadrant]
                row = 9 + quadrant
                
                ws.cell(row=row, column=1, value=f"象限{quadrant}")
                ws.cell(row=row, column=2, value=stats['name'])
                ws.cell(row=row, column=3, value=stats['description'])
                ws.cell(row=row, column=4, value=stats['count'])
                ws.cell(row=row, column=5, value=stats['strategy'])
                
                # 设置边框
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = self.border
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 50
    
    def _create_charts_sheet(self) -> None:
        """创建可视化图表工作表"""
        ws = self.workbook.create_sheet("可视化图表 (Charts)")
        
        ws['A1'] = "可视化图表"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_alignment
        ws.merge_cells('A1:D1')
        
        # 说明文字
        ws['A3'] = "注意: 此工作表用于存放分析生成的图表图片"
        ws['A4'] = "在实际应用中，可以通过程序将ECharts生成的图表保存为图片并插入到此处"
        
        # 图表占位符
        chart_types = [
            "四象限散点图",
            "帕累托曲线图", 
            "分布区间图",
            "盈亏分析图",
            "贡献度分析图"
        ]
        
        for i, chart_type in enumerate(chart_types, 6):
            ws[f'A{i}'] = f"{i-5}. {chart_type}"
            ws[f'A{i}'].font = self.header_font
    
    def _create_analyzed_data_sheet(self) -> None:
        """创建分析结果数据工作表"""
        ws = self.workbook.create_sheet("分析结果数据 (Analyzed Data)")
        
        analysis_result = self.analysis_data.get('analysis_result', {})
        aggregated_data = analysis_result.get('aggregated_data', [])
        
        if not aggregated_data:
            ws['A1'] = "无分析数据"
            return
        
        # 转换为DataFrame
        df = pd.DataFrame(aggregated_data)
        
        # 添加象限信息（如果存在）
        quadrant_analysis = analysis_result.get('quadrant_analysis', {})
        scatter_data = quadrant_analysis.get('scatter_data', [])
        
        if scatter_data:
            scatter_df = pd.DataFrame(scatter_data)
            # 合并象限信息
            if len(scatter_df) == len(df):
                df['象限归类'] = scatter_df.get('象限名称', '')
                df['建议策略'] = scatter_df.get('建议策略', '')
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # 设置数据边框
        for row in ws.iter_rows(min_row=2, max_row=len(df)+1, max_col=len(df.columns)):
            for cell in row:
                cell.border = self.border
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_raw_data_sheet(self) -> None:
        """创建原始数据工作表"""
        ws = self.workbook.create_sheet("原始数据 (含标注) (Raw Data with Tags)")
        
        analysis_result = self.analysis_data.get('analysis_result', {})
        processed_data = analysis_result.get('processed_data', [])
        
        if not processed_data:
            ws['A1'] = "无原始数据"
            return
        
        # 转换为DataFrame
        df = pd.DataFrame(processed_data)
        
        # 添加象限标注
        # 这里需要根据原始数据的分组字段来匹配象限信息
        analysis_type = analysis_result.get('analysis_type', '')
        quadrant_analysis = analysis_result.get('quadrant_analysis', {})
        scatter_data = quadrant_analysis.get('scatter_data', [])
        
        if scatter_data and analysis_type:
            # 创建象限映射字典
            quadrant_map = {}
            field_mapping = analysis_result.get('field_detection', {}).get('detected_fields', {})
            
            group_field_map = {
                'product': 'product',
                'customer': 'customer',
                'region': 'region'
            }
            
            group_field = group_field_map.get(analysis_type)
            if group_field and group_field in field_mapping:
                group_column = field_mapping[group_field]
                
                for item in scatter_data:
                    if group_column in item:
                        quadrant_map[item[group_column]] = item.get('象限名称', '')
                
                # 为原始数据添加象限标注
                if group_column in df.columns:
                    df['象限归类'] = df[group_column].map(quadrant_map).fillna('未分类')
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # 设置数据边框
        for row in ws.iter_rows(min_row=2, max_row=len(df)+1, max_col=len(df.columns)):
            for cell in row:
                cell.border = self.border
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
